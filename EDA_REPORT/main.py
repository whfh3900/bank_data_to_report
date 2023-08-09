# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
import datetime
import pandas as pd
import numpy as np
import os
import pytz
import time
# import argparse
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from DCP_Understanding import SetDataFrame
from DCP_Visualization import SetGraph
from DCP_Utils import *
from konfig import Config
cc = Config("./conf.ini").get_map("report")
KST = pytz.timezone('Asia/Seoul')
datetime.datetime.now(KST)


# # ArgumentParser 객체 생성
# parser = argparse.ArgumentParser(description="An example argument parser")

# # 인자 추가
# parser.add_argument('--path', '-p', default="./data" ,help='The path to the folder where the data files are located')
# parser.add_argument('--data', '-d', default="account_trans", help='Table or data file name to analyze')
# parser.add_argument('--company', '-c', default="닉컴퍼니", help='Company name to which it belongs')
# parser.add_argument('--user', '-u', default="NIC", help='User name')
# parser.add_argument('--output', '-o', default="./result", help='Result Report File Path')

# # 명령행 인자 파싱
# args = parser.parse_args()

#### input
# path = args.path
# table = args.data
# result_report_file_path = os.path.join(args.output, "%s_report_example.csv"%table)
# company = args.company
# user = args.user


######### 표지
def Cover(sheet1):
    img = Image('./EDA_REPORT/img/eda_report.png')
    sheet1.add_image(img, "C6")

    sheet1['D21'] = '회사명'
    sheet1['D22'] = '데이터명'
    sheet1['D23'] = '작성일자'
    sheet1['D24'] = '작성시간'
    sheet1['D25'] = '작성자'

    sheet1['E21'] = company
    sheet1['E22'] = data
    sheet1['E23'] = today
    sheet1['E24'] = nowtime
    sheet1['E25'] = user


    for j in ['D', 'E']:
        max_length = 0
        for i in range(21, 26):
            sheet1['{}{}'.format(j, i)].border = border_styles()
            sheet1['{}{}'.format(j, i)].font = font_styles()
            if j == 'D':
                sheet1['{}{}'.format(j, i)].fill = patternfill_styles(23)
            elif j == 'E':
                sheet1['{}{}'.format(j, i)].fill = patternfill_styles(22)

            if len(str(sheet1['{}{}'.format(j, i)].value)) > max_length:
                max_length = len(str(sheet1['{}{}'.format(j, i)].value))
                
        adjusted_width = (max_length + 2) * 1.5
        sheet1.column_dimensions[j].width = adjusted_width


## 데이터 이해
def Understanding_Data(sheet2):
    un = SetDataFrame(df, bank, stat, code)

    shape = un.dcp_shape()
    types = un.dcp_types()
    missing = un.dcp_missing()
    unique = un.dcp_unique()
    outliers = un.dcp_outliers()

    sheet2['C2'] = '행'
    sheet2['D2'] = '열'
    sheet2['B3'] = '데이터 크기'
    sheet2['C3'] = shape[0]
    sheet2['D3'] = shape[1]

    for j in ['B', 'C', 'D']:
        for i in range(2, 4):
            sheet2['{}{}'.format(j, i)].border = border_styles()
            sheet2['{}{}'.format(j, i)].font = font_styles()
            sheet2['{}{}'.format(j, i)].alignment = alignment_styles()
            if (j != 'B') and (i==3):
                sheet2['{}{}'.format(j, i)].fill = patternfill_styles(22)
            else:
                sheet2['{}{}'.format(j, i)].fill = patternfill_styles(23)

    sheet2['B7'] = '필드명'
    sheet2['C7'] = '데이터타입'
    sheet2['D7'] = '결측값개수'
    sheet2['E7'] = '클래스개수'
    sheet2['F7'] = '이상치개수'

    r = 8
    for i,n in enumerate(df.columns):
        sheet2['B{}'.format(r)] = df.columns[i]
        sheet2['C{}'.format(r)] = types[n]
        sheet2['D{}'.format(r)] = len(missing[n])
        sheet2['E{}'.format(r)] = len(unique[n])
        sheet2['F{}'.format(r)] = len(outliers[n])
        r+=1

    for j in ['B', 'C', 'D', 'E', 'F']:
        max_length = 0
        for i in range(7, r):
            sheet2['{}{}'.format(j, i)].alignment = alignment_styles()
            sheet2['{}{}'.format(j, i)].border = border_styles()
            if i == 7:
                sheet2['{}{}'.format(j, i)].font = font_styles()
                sheet2['{}{}'.format(j, i)].fill = patternfill_styles(23)
            elif (i != 7) and (j == 'B'):
                sheet2['{}{}'.format(j, i)].font = font_styles()
                sheet2['{}{}'.format(j, i)].fill = patternfill_styles(22)
                
            if len(str(sheet2['{}{}'.format(j, i)].value)) > max_length:
                max_length = len(str(sheet2['{}{}'.format(j, i)].value))
                
        adjusted_width = (max_length + 2) * 1.5
        sheet2.column_dimensions[j].width = adjusted_width

    return missing, outliers



######### 결측치 데이터
def Miss_Data(sheet3, missing):
    for n in tqdm(df.columns):
        if sum([len(i) for i in missing.values()]) == 0:
            empty_list = ['']*len(df.columns)
            empty_list[0] = '결측값 없음'
            sheet3.append(empty_list)
            break
        else:
            missing_data = df[df.index.isin(missing[n])].reset_index()
            if (len(missing_data) == 0) | (len(missing_data) == len(df)):
                pass
            else:
                empty_list = ['']*len(df.columns)
                empty_list[0] = '결측 컬럼 :'
                empty_list[1] = n
                sheet3.append(empty_list)
                sheet3.append(list(missing_data.columns))
                for r in dataframe_to_rows(missing_data, index=False, header=False):
                    sheet3.append(r)
                sheet3.append(['']*len(df.columns))
                sheet3.append(['']*len(df.columns))



######### 이상치 데이터
def Outlier_Data(sheet4, outliers):
    for n in tqdm(df.columns):
        if sum([len(i) for i in outliers.values()]) == 0:
            empty_list = ['']*len(df.columns)
            empty_list[0] = '이상치 없음'
            sheet4.append(empty_list)
            break
        else:
            outliers_data = df[df.index.isin(outliers[n])].reset_index()
            if (len(outliers_data) == 0) | (len(outliers_data) == len(df)):
                pass
            else:
                empty_list = ['']*len(df.columns)
                empty_list[0] = n
                empty_list[1] = ': 정해진 타입과 형식이 맞지 않음.'
                sheet4.append(empty_list)
                sheet4.append(list(outliers_data.columns))
                for r in dataframe_to_rows(outliers_data, index=False, header=False):
                    sheet4.append(r)
                sheet4.append(['']*len(df.columns))
                sheet4.append(['']*len(df.columns))




######### 명세서
def Spec(sheet7):
    ######### 명세서
    stat_col = ['은행명','항목명','항목의미','항목형식',\
        '설명','사용코드','비고']
    sheet7.append(stat_col)
    for r in dataframe_to_rows(stat, index=False, header=False):
        sheet7.append(r)


    ######### 코드표
def Code(sheet8):
    code_col = ['은행명','사용코드','사용변수','변수설명','비고']
    sheet8.append(code_col)
    for r in dataframe_to_rows(code, index=False, header=False):
        sheet8.append(r)



if __name__ == "__main__":
    
    path = cc["path"]                                               ## 기본경로
    bank = cc["bank"]                                               ## 데이터 출처 은행
    data = cc["data"]                                               ## 데이터 파일명

    result_report_file_path = \
        os.path.join(cc["output"], "%s_report_example.csv"%data)    ## 결과보고서 경로
    company = cc["company"]                                         ## 사용자 회사
    user = cc["user"]                                               ## 사용자명

    today = datetime.datetime.now().strftime('%Y-%m-%d')
    nowtime = datetime.datetime.now().strftime('%H:%M:%S')

    ### 상태표와 코드표
    stat = pd.read_csv(cc["stat_file"], encoding="cp949")           ## 데이터 상태표 파일경로
    code = pd.read_csv(cc["code_file"], encoding="cp949")           ## 데이터 코드표 파일경로

    ### 데이터
    df = pd.read_csv(os.path.join(path, data+".csv"), \
        encoding="cp949")
    
    ### 그래프
    pltpath = cc["save_graph_image"]                                ## 그래프 생성 경로
    sg = SetGraph(df, bank, stat, pltpath)
    
    
    ######### 시트생성
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = '표지'
    sheet2 = wb.create_sheet('데이터 이해',2)
    sheet3 = wb.create_sheet('결측치 데이터',3)
    sheet4 = wb.create_sheet('이상치 데이터',4)
    sheet5 = wb.create_sheet('분포도',5)    
    sheet6 = wb.create_sheet('스케일',6)
    sheet7 = wb.create_sheet('명세서',7)
    sheet8 = wb.create_sheet('코드표',8)
    
    ######### 보고서 작성실행
    Cover(sheet1)                                                   ## 표지                                                   
    missing, outliers = Understanding_Data(sheet2)                  ## 데이터 이해
    Miss_Data(sheet3, missing)                                      ## 결측 데이터
    Outlier_Data(sheet4, outliers)                                  ## 이상치 데이터
    sg.dcp_distribution_graph(sheet5)                               ## 범주형 분포도 그래프
    sg.dcp_scale_comparison(sheet6)                                 ## 수치형 스케일 그래프
    Spec(sheet7)                                                    ## 데이터 명세서
    Code(sheet8)                                                    ## 데이터 코드표
    wb.save(result_report_file_path)                                ## 보고서 저장
